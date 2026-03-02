"""
Backend Recensement BAT-EQ-127 — Upload PDF et remplissage feuille Recensement.
Architecture client-serveur : API FastAPI + interface web.
"""
import io
import logging
import os
import re
import uuid
from datetime import datetime, timedelta
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

from pydantic import BaseModel
from fastapi import FastAPI, File, UploadFile, HTTPException, Request, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse

# Dossier du projet (parent du backend) — tous les fichiers téléversés sont sauvegardés ici
PROJECT_ROOT = Path(__file__).resolve().parent.parent
UPLOADS_DIR = PROJECT_ROOT / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
SPLITS_DIR = UPLOADS_DIR / "splits"
SPLITS_DIR.mkdir(exist_ok=True)


def _safe_upload_filename(name: str) -> str:
    """Nom de fichier sûr pour le stockage (sans chemin, caractères dangereux remplacés)."""
    if not name or not name.strip():
        return f"document_{uuid.uuid4().hex[:8]}.pdf"
    base = Path(name).name.strip()
    base = re.sub(r'[<>:"/\\|?*]', "_", base)
    return base or f"document_{uuid.uuid4().hex[:8]}.pdf"

# Taille max par fichier (50 Mo)
MAX_FILE_SIZE = 50 * 1024 * 1024
ALLOWED_EXTENSIONS = {".pdf"}

app = FastAPI(
    title="Recensement BAT-EQ-127",
    description="Upload de PDF pour alimenter le tableau de recensement",
    version="1.0.0",
)


@app.get("/api/health")
def health():
    return {"status": "ok", "message": "Recensement BAT-EQ-127 API"}


@app.post("/api/upload")
async def upload_pdf(
    file: UploadFile = File(...),
    overwrite: bool = Query(False, description="Écraser si le fichier existe déjà"),
):
    """Reçoit un ou plusieurs PDF ; pour l’instant un seul fichier par requête."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="Nom de fichier manquant")
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Seuls les fichiers PDF sont acceptés.",
        )
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail="Fichier trop volumineux (max 50 Mo).",
        )
    safe_name = _safe_upload_filename(file.filename)
    if not safe_name.lower().endswith(".pdf"):
        safe_name = safe_name + ".pdf"
    dest = UPLOADS_DIR / safe_name
    if dest.is_file() and not overwrite:
        raise HTTPException(
            status_code=409,
            detail="Un fichier avec ce nom existe déjà. Confirmez pour l'écraser.",
            headers={"X-Existing-File": safe_name},
        )
    with open(dest, "wb") as f:
        f.write(content)
    return {
        "success": True,
        "filename": file.filename,
        "saved_as": safe_name,
        "size": len(content),
    }


@app.post("/api/upload-multiple")
async def upload_multiple(
    files: list[UploadFile] = File(...),
    overwrite: bool = Query(False, description="Écraser les fichiers déjà existants"),
):
    """Reçoit plusieurs PDF. Si un fichier du même nom existe et overwrite=False, renvoie 409."""
    if not files:
        raise HTTPException(status_code=400, detail="Aucun fichier envoyé")
    existing = []
    for file in files:
        if not file.filename:
            continue
        safe_name = _safe_upload_filename(file.filename)
        if not safe_name.lower().endswith(".pdf"):
            safe_name = safe_name + ".pdf"
        dest = UPLOADS_DIR / safe_name
        if dest.is_file() and not overwrite:
            existing.append(safe_name)
    if existing:
        raise HTTPException(
            status_code=409,
            detail={"message": "Un ou plusieurs fichiers existent déjà. Confirmez pour les écraser.", "existing": existing},
        )
    results = []
    for file in files:
        if not file.filename:
            results.append({"filename": None, "success": False, "error": "Nom manquant"})
            continue
        ext = Path(file.filename).suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            results.append({
                "filename": file.filename,
                "success": False,
                "error": "Seuls les PDF sont acceptés",
            })
            continue
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            results.append({
                "filename": file.filename,
                "success": False,
                "error": "Fichier trop volumineux (max 50 Mo)",
            })
            continue
        safe_name = _safe_upload_filename(file.filename)
        if not safe_name.lower().endswith(".pdf"):
            safe_name = safe_name + ".pdf"
        dest = UPLOADS_DIR / safe_name
        with open(dest, "wb") as f:
            f.write(content)
        results.append({
            "filename": file.filename,
            "success": True,
            "saved_as": safe_name,
            "size": len(content),
        })
    return {"results": results}


@app.get("/api/uploads")
def list_uploads():
    """Liste les PDF déjà téléversés (hors sous-dossier splits)."""
    files = []
    for p in UPLOADS_DIR.iterdir():
        if p.is_file() and p.suffix.lower() == ".pdf":
            files.append({
                "name": p.name,
                "size": p.stat().st_size,
                "upload_time": p.stat().st_mtime,
            })
    files.sort(key=lambda x: x["upload_time"], reverse=True)
    return {"files": files}


def _sanitize_split_stem(name: str) -> str:
    """Nom de dossier sûr pour Windows : espaces en fin/début supprimés, caractères invalides remplacés."""
    if not name:
        return "splits"
    s = str(name).strip()
    # Caractères interdits sous Windows : \ / : * ? " < > |
    for c in r'\/:*?"<>|':
        s = s.replace(c, "_")
    return s or "splits"


class SplitDevisRequest(BaseModel):
    filename: str


@app.post("/api/split-devis")
def split_devis(body: SplitDevisRequest):
    """
    Découpe un PDF déjà téléversé par numéro de devis (OCR Tesseract en haut de page).
    Retourne la liste des PDF créés (un par devis).
    """
    filename = body.filename
    if not filename or ".." in filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Nom de fichier invalide")
    pdf_path = UPLOADS_DIR / filename
    if not pdf_path.is_file():
        raise HTTPException(status_code=404, detail="PDF introuvable")
    safe_stem = _sanitize_split_stem(pdf_path.stem)
    out_subdir = SPLITS_DIR / safe_stem
    try:
        from backend.pdf_split_by_devis import split_pdf_by_devis
    except ImportError:
        from pdf_split_by_devis import split_pdf_by_devis
    try:
        results = split_pdf_by_devis(pdf_path, out_subdir)
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail="Erreur découpage : " + str(e))
    base_url = f"/api/splits/{safe_stem}"
    return {
        "source": filename,
        "source_stem": safe_stem,
        "devis_count": len(results),
        "devis": [
            {
                "devis_number": r["devis_number"],
                "filename": r["filename"],
                "page_count": r["page_count"],
                "download_url": f"{base_url}/{r['filename']}",
            }
            for r in results
        ],
    }


@app.get("/api/splits/{source_stem}/{out_filename:path}")
def download_split(request: Request, source_stem: str, out_filename: str):
    """Téléchargement d’un PDF découpé."""
    if ".." in source_stem or ".." in str(out_filename):
        raise HTTPException(status_code=400, detail="Chemin invalide")
    path = SPLITS_DIR / source_stem / out_filename
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Fichier introuvable")
    inline = request.query_params.get("display") == "1"
    disposition = "inline" if inline else "attachment"
    return FileResponse(
        path,
        filename=path.name,
        media_type="application/pdf",
        headers={"Content-Disposition": f'{disposition}; filename="{path.name}"'},
    )


class RetraiterDevisRequest(BaseModel):
    source_stem: str
    current_filename: str


class SupprimerDevisRequest(BaseModel):
    source_stem: str
    filename: str


def _pdf_page_count(pdf_path: Path) -> int:
    """Retourne le nombre de pages d'un PDF (PyMuPDF)."""
    try:
        import fitz
    except ImportError:
        return 0
    try:
        doc = fitz.open(pdf_path)
        n = len(doc)
        doc.close()
        return n
    except Exception:
        return 0


def _list_pdf_in_folder(folder: Path) -> list[str]:
    """Liste uniquement les .pdf (pas .tmp) pour les logs."""
    return sorted(p.name for p in folder.iterdir() if p.is_file() and p.suffix.lower() == ".pdf")


def _list_all_files_in_folder(folder: Path) -> list[str]:
    """Liste tous les fichiers (pdf, tmp, etc.) pour debug duplication."""
    return sorted(p.name for p in folder.iterdir() if p.is_file())


@app.post("/api/split-devis/retraiter")
def retraiter_devis(body: RetraiterDevisRequest):
    """
    Déplace la dernière page du PDF suivant vers le début du PDF
    sur lequel on a cliqué « Retraiter », et la supprime du PDF suivant.
    """
    if not body.source_stem or ".." in body.source_stem or "/" in body.source_stem or "\\" in body.source_stem:
        raise HTTPException(status_code=400, detail="Source invalide")
    if not body.current_filename or ".." in body.current_filename or "/" in body.current_filename or "\\" in body.current_filename:
        raise HTTPException(status_code=400, detail="Nom de fichier invalide")
    folder = SPLITS_DIR / body.source_stem
    if not folder.is_dir():
        raise HTTPException(status_code=404, detail="Dossier source introuvable")
    try:
        import fitz
    except ImportError:
        raise HTTPException(status_code=500, detail="PyMuPDF (fitz) est requis pour le retraitement")

    pdf_files = _list_pdf_in_folder(folder)
    log.info("[RETRAITER] DEBUT source_stem=%s current_filename=%s nb_pdf_avant=%d", body.source_stem, body.current_filename, len(pdf_files))
    log.info("[RETRAITER] Liste PDF avant: %s", pdf_files)
    log.info("[RETRAITER] Tous fichiers dans dossier avant: %s", _list_all_files_in_folder(folder))

    if body.current_filename not in pdf_files:
        raise HTTPException(status_code=404, detail="Fichier introuvable dans cette source")
    idx = pdf_files.index(body.current_filename)
    if idx >= len(pdf_files) - 1:
        raise HTTPException(status_code=400, detail="Aucun PDF suivant : le dernier fichier ne peut pas être retraité")
    next_filename = pdf_files[idx + 1]
    next_path = folder / next_filename
    curr_path = folder / body.current_filename
    log.info("[RETRAITER] Fichiers cibles: curr=%s next=%s", body.current_filename, next_filename)

    doc_curr = fitz.open(curr_path)
    doc_next = fitz.open(next_path)
    n_next = len(doc_next)
    if n_next == 0:
        doc_curr.close()
        doc_next.close()
        raise HTTPException(status_code=400, detail="Le PDF suivant n'a plus de pages")

    def _overwrite_pdf(target: Path, content: bytes, label: str) -> None:
        tmp = target.with_suffix(target.suffix + ".tmp")
        log.info("[RETRAITER] %s: target=%s tmp=%s", label, target.name, tmp.name)
        try:
            tmp.write_bytes(content)
            log.info("[RETRAITER] %s: write_bytes(tmp) ok len=%d", label, len(content))
            target.unlink(missing_ok=True)
            log.info("[RETRAITER] %s: unlink(target) ok", label)
            tmp.rename(target)
            log.info("[RETRAITER] %s: rename(tmp->target) ok", label)
        except Exception as e:
            log.exception("[RETRAITER] %s: ERREUR %s", label, e)
            raise
        finally:
            if tmp.exists():
                try:
                    tmp.unlink()
                    log.info("[RETRAITER] %s: tmp encore present -> unlink(tmp) ok", label)
                except Exception as e2:
                    log.warning("[RETRAITER] %s: unlink(tmp) fail: %s", label, e2)

    try:
        # 1) Nouveau contenu du PDF courant : dernière page du next en première position, puis toutes les pages du curr
        new_curr = fitz.open()
        new_curr.insert_pdf(doc_next, from_page=n_next - 1, to_page=n_next - 1)
        new_curr.insert_pdf(doc_curr, from_page=0, to_page=len(doc_curr) - 1)
        doc_curr.close()
        doc_curr = None
        buf_curr = io.BytesIO()
        new_curr.save(buf_curr)
        new_curr.close()
        _overwrite_pdf(curr_path, buf_curr.getvalue(), "CURR")
        apres_curr = _list_pdf_in_folder(folder)
        log.info("[RETRAITER] Apres ecrasement CURR: nb_pdf=%d liste=%s", len(apres_curr), apres_curr)
        log.info("[RETRAITER] Tous fichiers apres CURR: %s", _list_all_files_in_folder(folder))
    except Exception as e:
        if doc_curr is not None:
            try:
                doc_curr.close()
            except Exception:
                pass
        try:
            doc_next.close()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail="Erreur écriture PDF courant : " + str(e))
    try:
        # 2) PDF suivant : supprimer si 1 page, sinon écraser sans la dernière page
        if n_next == 1:
            doc_next.close()
            next_path.unlink(missing_ok=True)
            log.info("[RETRAITER] NEXT: 1 page -> unlink(next) ok")
        else:
            new_next = fitz.open()
            new_next.insert_pdf(doc_next, from_page=0, to_page=n_next - 2)
            doc_next.close()
            doc_next = None
            buf_next = io.BytesIO()
            new_next.save(buf_next)
            new_next.close()
            _overwrite_pdf(next_path, buf_next.getvalue(), "NEXT")
        apres_next = _list_pdf_in_folder(folder)
        log.info("[RETRAITER] Apres NEXT: nb_pdf=%d liste=%s", len(apres_next), apres_next)
        log.info("[RETRAITER] Tous fichiers apres NEXT: %s", _list_all_files_in_folder(folder))
    except Exception as e:
        raise HTTPException(status_code=500, detail="Erreur écriture PDF suivant : " + str(e))

    base_url = f"/api/splits/{body.source_stem}"
    devis = []
    for name in _list_pdf_in_folder(folder):
        page_count = _pdf_page_count(folder / name)
        devis.append({
            "filename": name,
            "page_count": page_count,
            "download_url": f"{base_url}/{name}",
        })
    log.info("[RETRAITER] FIN nb_pdf_apres=%d (attendu: nb_avant-1 si next 1 page, sinon nb_avant)", len(devis))
    log.info("[RETRAITER] Tous fichiers en fin: %s", _list_all_files_in_folder(folder))
    return {"devis": devis}


@app.post("/api/split-devis/supprimer")
def supprimer_devis(body: SupprimerDevisRequest):
    """Supprime un PDF découpé (un devis) de la source."""
    if not body.source_stem or ".." in body.source_stem or "/" in body.source_stem or "\\" in body.source_stem:
        raise HTTPException(status_code=400, detail="Source invalide")
    if not body.filename or ".." in body.filename or "/" in body.filename or "\\" in body.filename:
        raise HTTPException(status_code=400, detail="Nom de fichier invalide")
    folder = SPLITS_DIR / body.source_stem
    if not folder.is_dir():
        raise HTTPException(status_code=404, detail="Dossier source introuvable")
    path = folder / body.filename
    if not path.is_file():
        raise HTTPException(status_code=404, detail="Fichier introuvable")
    try:
        path.unlink()
    except OSError as e:
        raise HTTPException(status_code=500, detail="Impossible de supprimer le fichier : " + str(e))
    base_url = f"/api/splits/{body.source_stem}"
    devis = []
    for name in _list_pdf_in_folder(folder):
        page_count = _pdf_page_count(folder / name)
        devis.append({
            "filename": name,
            "page_count": page_count,
            "download_url": f"{base_url}/{name}",
        })
    return {"devis": devis}


@app.get("/api/splits/sources")
def list_split_sources():
    """Liste les sources de PDF découpés (sous-dossiers de splits/)."""
    sources = []
    for p in SPLITS_DIR.iterdir():
        if p.is_dir():
            count = sum(1 for c in p.iterdir() if c.is_file() and c.suffix.lower() == ".pdf")
            if count:
                sources.append({"source_stem": p.name, "file_count": count})
    sources.sort(key=lambda x: x["source_stem"])
    return {"sources": sources}


@app.get("/api/splits/files")
def list_split_files(source: str):
    """Liste les PDF d'une source découpée."""
    if not source or ".." in source or "/" in source or "\\" in source:
        raise HTTPException(status_code=400, detail="Source invalide")
    folder = SPLITS_DIR / source
    if not folder.is_dir():
        return {"files": []}
    files = []
    for p in folder.iterdir():
        if p.is_file() and p.suffix.lower() == ".pdf":
            files.append({"name": p.name})
    files.sort(key=lambda x: x["name"])
    return {"files": files}


class ExtractDevisRequest(BaseModel):
    source_stem: str
    filenames: list[str] | None = None


@app.post("/api/extract-devis")
def extract_devis(body: ExtractDevisRequest):
    """Extrait les données de un ou plusieurs PDF devis (découpés)."""
    if not body.source_stem or ".." in body.source_stem:
        raise HTTPException(status_code=400, detail="Source invalide")
    folder = SPLITS_DIR / body.source_stem
    if not folder.is_dir():
        raise HTTPException(status_code=404, detail="Source introuvable")
    try:
        from backend.extract_devis_data import extract_devis_data
    except ImportError:
        from extract_devis_data import extract_devis_data
    pdf_files = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"]
    if body.filenames:
        pdf_files = [p for p in pdf_files if p.name in body.filenames]
    pdf_files.sort(key=lambda p: p.name)
    results = []
    for pdf_path in pdf_files:
        try:
            data = extract_devis_data(pdf_path)
            data["_filename"] = pdf_path.name
            results.append(data)
        except Exception as e:
            results.append({"_filename": pdf_path.name, "num_devis": "", "error": str(e)})
    return {"results": results}


class ExportExcelRequest(BaseModel):
    records: list[dict]
    save_path: str | None = None  # chemin relatif au projet pour enregistrer (ex. exports/recensement.xlsx)


def _num(s):
    """Convertit en nombre (pas d'espace, virgule → point). Conforme Lisez-moi : pas d'espace dans les nombres."""
    if s is None or s == "":
        return None
    if isinstance(s, (int, float)):
        return s
    s = str(s).replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _str_clean(s, code_postal=False):
    """Chaîne sans espaces avant/après. Lisez-moi : pas d'espace avant/après une chaîne ; pas d'espace dans le code postal."""
    if s is None:
        return None
    s = str(s).strip()
    if code_postal:
        s = s.replace(" ", "")
    return s if s else None


def _siren_from_siret(s):
    """SIREN = 9 premiers chiffres du SIRET client (sans espaces). Retourne None si vide."""
    if s is None:
        return None
    digits = "".join(c for c in str(s) if c.isdigit())
    return digits[:9] if digits else None


def _tel_digits(s):
    """Téléphone : uniquement les chiffres (sans tirets, espaces, etc.)."""
    if s is None:
        return None
    digits = "".join(c for c in str(s) if c.isdigit())
    return digits if digits else None


def _date_rai_fallback(row_index: int):
    """Si DATE RAI vide : 4 devis par jour entre 13/11/2025 et 28/12/2025."""
    base = datetime(2025, 11, 13)
    end = datetime(2025, 12, 28)
    day_offset = row_index // 4
    d = base + timedelta(days=day_offset)
    if d > end:
        d = end
    return d.strftime("%d/%m/%Y")


def _date_plus_15(date_str):
    """DATE D'ENGAGEMENT = DATE d'envoi du RAI + 15 jours. Retourne une chaîne ou None."""
    if not date_str or not str(date_str).strip():
        return None
    s = str(date_str).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            dt = datetime.strptime(re.sub(r"\s+", " ", s)[:10], fmt)
            if dt.year < 100:
                dt = dt.replace(year=dt.year + 2000)
            return (dt + timedelta(days=15)).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return None


def _find_recensement_template():
    """Retourne le chemin du modèle Excel Recensement (BONModèle prioritaire si présent)."""
    # Priorité : fichier BON Modèle BAT-EQ-127 (nouveau format, ligne 2 = en-têtes, données à partir ligne 4)
    for p in PROJECT_ROOT.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() != ".xlsx":
            continue
        name = p.name
        if "BAT" in name and "recensement" in name.lower():
            return p, "bon"  # BON template : start_row=4, colonnes 2-43
    # Sinon : exemple.xlsx (ancien format, données à partir ligne 2, colonnes 1-42)
    ex = PROJECT_ROOT / "exemple.xlsx"
    if ex.is_file():
        return ex, "exemple"
    return None, None


def _safe_export_path(relative_path: str) -> Path:
    """Chemin d'enregistrement sécurisé : relatif à PROJECT_ROOT, sans '..'."""
    clean = relative_path.strip().replace("\\", "/").lstrip("/")
    if ".." in clean or ":" in clean:
        raise HTTPException(status_code=400, detail="Chemin non autorisé")
    dest = (PROJECT_ROOT / clean).resolve()
    try:
        dest.relative_to(PROJECT_ROOT.resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="Chemin non autorisé")
    return dest


@app.post("/api/export-excel")
def export_excel(body: ExportExcelRequest):
    """Génère un Excel et soit le renvoie (téléchargement), soit l'enregistre à save_path."""
    if not body.records:
        raise HTTPException(status_code=400, detail="Aucun enregistrement")
    try:
        temp_path = _do_export_excel(body)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="Export Excel : " + str(e))
    if body.save_path:
        import os
        import shutil
        dest = _safe_export_path(body.save_path)
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(temp_path, dest)
        try:
            os.unlink(temp_path)
        except OSError:
            pass
        return JSONResponse(content={"saved": True, "path": body.save_path})
    return FileResponse(
        temp_path,
        filename="recensement_bat_eq_127.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# En-têtes exacts pour l'export Excel (une seule ligne, 37 colonnes) — on n'utilise pas les en-têtes du BON modèle
EXCEL_HEADERS = (
    "Opération n°",
    "Type de trame",
    "Code Fiche",
    "RAISON SOCIALE du demandeur",
    "SIREN du demandeur",
    "REFERENCE PIXEL",
    "DATE d'envoi du RAI",
    "MONTANT de l'incitation financière CEE",
    "DATE D'ENGAGEMENT",
    "Raison sociale du mandataire assurant le rôle actif et incitatif",
    "Numéro SIREN du mandataire assurant le rôle actif et incitatif",
    "Nature de la bonification",
    "NOM du bénéficiaire",
    "PRENOM du bénéficiaire",
    "ADRESSE de l'opération",
    "CODE POSTAL",
    "VILLE",
    "Numéro de téléphone du bénéficiaire",
    "Adresse de courriel du bénéficiaire",
    "NOM DU SITE bénéficiaire",
    "ADRESSE de l'opération (site)",
    "CODE POSTAL (site)",
    "VILLE (site)",
    "Numéro de téléphone (site)",
    "Adresse de courriel (site)",
    "SIREN du professionnel mettant en œuvre l'opération",
    "RAISON SOCIALE du professionnel mettant en œuvre",
    "NUMERO de devis",
    "Numéro Client",
    "MONTANT du devis (€ TTC)",
    "RAISON SOCIALE du professionnel figurant sur le devis",
    "SIREN du professionnel figurant sur le devis",
    "Nombre de luminaires",
    "Puissance totale des luminaires à modules LED (W)",
    "Puissance des luminaires LED avec IRC < 90 (W)",
    "Puissance des luminaires LED avec IRC ≥ 90 et R9 > 0 (W)",
    "Secteur concerné",
    "Précision sur le secteur concerné",
)


def _do_export_excel(body: ExportExcelRequest):
    """Crée un NOUVEAU classeur Excel (une seule feuille), sans aucun fichier modèle."""
    try:
        import openpyxl
        import tempfile
        import os
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl requis")
    wb = openpyxl.Workbook()
    # Une seule feuille : garder la feuille active et la renommer "Recensement"
    ws = wb.active
    ws.title = "Recensement"
    max_col = 38
    header_row = 1
    start_row = 2
    num_format_cols = (8, 30, 34, 35)
    format_sans_espace = "0,00"

    # Ligne 1 : en-têtes
    for col_idx, label in enumerate(EXCEL_HEADERS, start=1):
        ws.cell(row=header_row, column=col_idx, value=label)

    def col(j):
        return j

    for i, rec in enumerate(body.records):
        rec = rec if isinstance(rec, dict) else {}
        row = start_row + i
        date_rai_raw = _str_clean(rec.get("date_rai"))
        date_rai = date_rai_raw or _date_rai_fallback(i)
        date_engagement = _date_plus_15(date_rai_raw or date_rai)
        tel_digits = _tel_digits(rec.get("tel"))
        ws.cell(row=row, column=col(2), value="TPM")
        ws.cell(row=row, column=col(3), value=_str_clean(rec.get("code_fiche") or "BAT-EQ-127"))
        ws.cell(row=row, column=col(4), value=_str_clean(rec.get("nom_site_beneficiaire")))
        ws.cell(row=row, column=col(5), value=_str_clean(_siren_from_siret(rec.get("siret"))))
        ws.cell(row=row, column=col(7), value=date_rai)
        ws.cell(row=row, column=col(8), value=_num(rec.get("prime_cee")))
        ws.cell(row=row, column=col(9), value=date_engagement)
        ws.cell(row=row, column=col(10), value="OTC FLOW France")
        ws.cell(row=row, column=col(11), value="953658036")
        ws.cell(row=row, column=col(12), value="ZNI")
        ws.cell(row=row, column=col(13), value=_str_clean(rec.get("nom_beneficiaire")))
        ws.cell(row=row, column=col(14), value=_str_clean(rec.get("prenom_beneficiaire")))
        ws.cell(row=row, column=col(15), value=_str_clean(rec.get("adresse_des_travaux")))
        ws.cell(row=row, column=col(16), value=_str_clean(rec.get("code_postal_travaux") or rec.get("code_postal"), code_postal=True))
        ws.cell(row=row, column=col(17), value=_str_clean(rec.get("ville_travaux") or rec.get("ville")))
        ws.cell(row=row, column=col(18), value=tel_digits)
        ws.cell(row=row, column=col(19), value=_str_clean(rec.get("mail")))
        ws.cell(row=row, column=col(20), value=_str_clean(rec.get("nom_site_beneficiaire")))
        ws.cell(row=row, column=col(21), value=_str_clean(rec.get("adresse_client")))
        ws.cell(row=row, column=col(22), value=_str_clean(rec.get("code_postal"), code_postal=True))
        ws.cell(row=row, column=col(23), value=_str_clean(rec.get("ville")))
        ws.cell(row=row, column=col(24), value=tel_digits)
        ws.cell(row=row, column=col(25), value=_str_clean(rec.get("mail")))
        ws.cell(row=row, column=col(26), value="421747007")
        ws.cell(row=row, column=col(27), value="LECBA.TP SARL")
        ws.cell(row=row, column=col(28), value=_str_clean(rec.get("num_devis") or rec.get("num_client")))
        ws.cell(row=row, column=col(29), value=_str_clean(rec.get("num_client")))
        ws.cell(row=row, column=col(30), value=_num(rec.get("prime_cee")))
        ws.cell(row=row, column=col(31), value="LECBA.TP SARL")
        ws.cell(row=row, column=col(32), value="42174700700035")
        ws.cell(row=row, column=col(33), value=_str_clean(rec.get("nombre_luminaires")))
        nb_led = _num(rec.get("nombre_luminaires"))
        ws.cell(row=row, column=col(34), value=(nb_led * 250) if nb_led is not None else None)
        ws.cell(row=row, column=col(35), value=_num(rec.get("irc")))
        ws.cell(row=row, column=col(38), value=_str_clean(rec.get("secteur_activite")))
        for col_idx in num_format_cols:
            ws.cell(row=row, column=col_idx).number_format = format_sans_espace

    ws.freeze_panes = "A2"
    try:
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        return path
    except Exception as e:
        raise HTTPException(status_code=500, detail="Erreur lors de l'export Excel: " + str(e))


# --- Lecture des enregistrements (PDF + exports Excel) ---
EXPORTS_DIR = PROJECT_ROOT / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)


@app.get("/api/enregistrements")
def list_enregistrements():
    """Liste tout ce qui est enregistré : PDF téléversés et fichiers Excel exportés."""
    uploads = []
    for p in UPLOADS_DIR.iterdir():
        if p.is_file() and p.suffix.lower() == ".pdf":
            rel = p.relative_to(PROJECT_ROOT)
            uploads.append({
                "name": p.name,
                "path": str(rel).replace("\\", "/"),
                "size": p.stat().st_size,
                "modified": p.stat().st_mtime,
                "type": "pdf",
            })
    uploads.sort(key=lambda x: x["modified"], reverse=True)
    exports = []
    for p in EXPORTS_DIR.iterdir():
        if p.is_file() and p.suffix.lower() in (".xlsx", ".xls"):
            rel = p.relative_to(PROJECT_ROOT)
            exports.append({
                "name": p.name,
                "path": str(rel).replace("\\", "/"),
                "size": p.stat().st_size,
                "modified": p.stat().st_mtime,
                "type": "excel",
            })
    exports.sort(key=lambda x: x["modified"], reverse=True)
    return {"uploads": uploads, "exports": exports}


@app.get("/api/enregistrements/download")
def download_enregistrement(path: str = Query(..., description="Chemin relatif (ex. uploads/fichier.pdf ou exports/recensement.xlsx)")):
    """Télécharge un fichier enregistré (PDF ou Excel) par son chemin relatif au projet."""
    path_clean = path.strip().replace("\\", "/").lstrip("/")
    if not path_clean or ".." in path_clean:
        raise HTTPException(status_code=400, detail="Chemin invalide")
    dest = (PROJECT_ROOT / path_clean).resolve()
    try:
        dest.relative_to(PROJECT_ROOT.resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="Chemin non autorisé")
    if not dest.is_file():
        raise HTTPException(status_code=404, detail="Fichier introuvable")
    return FileResponse(
        dest,
        filename=dest.name,
        media_type="application/octet-stream",
    )


# Servir les fichiers statiques (frontend)
STATIC_DIR = PROJECT_ROOT / "static"
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


@app.get("/")
def index():
    """Page d'accueil : interface d'upload."""
    index_path = STATIC_DIR / "index.html"
    if index_path.exists():
        return FileResponse(index_path)
    return JSONResponse(
        content={"message": "Ajoutez le dossier static avec index.html"},
        status_code=404,
    )
